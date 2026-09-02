"""
Portal do Cliente — Logás
=========================
Serve a cada cliente externo APENAS as leituras dos pontos dele.

Como funciona
-------------
- Lê o mesmo `dados.js` que a Torre de Controle usa (nada é duplicado).
- `clientes.json` diz quais pontos pertencem a cada login.
- A filtragem acontece AQUI, no servidor. O navegador do cliente nunca
  recebe dado de outro cliente, nem frota, nem plano de despacho.
- Sessão em cookie assinado (HMAC), senha guardada como hash PBKDF2.

Rodar:
    uvicorn app:app --host 127.0.0.1 --port 8100
"""

from __future__ import annotations

import hashlib
import hmac
import json
import os
import re
import secrets
import time
from base64 import urlsafe_b64decode, urlsafe_b64encode
from datetime import datetime, timedelta, timezone
from pathlib import Path

from fastapi import FastAPI, Form, Request, Response
from fastapi.responses import (FileResponse, HTMLResponse, JSONResponse,
                               RedirectResponse)
from fastapi.staticfiles import StaticFiles

# --------------------------------------------------------------------------- #
# Configuração
# --------------------------------------------------------------------------- #

AQUI = Path(__file__).resolve().parent

CLIENTES_JSON = Path(os.environ.get("LOGAS_CLIENTES", AQUI / "clientes.json"))
CONFIG_JSON = Path(os.environ.get("LOGAS_CONFIG", AQUI / "config.json"))
ADMINS_JSON = Path(os.environ.get("LOGAS_ADMINS", AQUI / "admins.json"))

# Chave de assinatura da sessão. Ordem: variável de ambiente, arquivo secret.txt
# (criado na primeira execução), sorteio em memória. O arquivo existe para o
# Windows, onde é mais simples que variável de ambiente de serviço: sem ele,
# cada reinício derrubaria as sessões abertas.
def _segredo() -> bytes:
    do_ambiente = os.environ.get("LOGAS_SECRET", "").strip()
    if do_ambiente:
        return do_ambiente.encode()
    arq = AQUI / "secret.txt"
    try:
        if arq.exists():
            valor = arq.read_text(encoding="utf-8-sig").strip()
            if valor:
                return valor.encode()
        valor = secrets.token_hex(32)
        arq.write_text(valor, encoding="utf-8")
        try:
            arq.chmod(0o600)
        except OSError:
            pass
        return valor.encode()
    except OSError:
        return secrets.token_bytes(32)


SECRET = _segredo()

# Cookie só viaja em HTTPS. Deixe "0" apenas para testar em localhost.
COOKIE_SECURE = os.environ.get("LOGAS_COOKIE_SECURE", "1") != "0"
SESSAO_HORAS = float(os.environ.get("LOGAS_SESSAO_HORAS", "12"))

TZ = timezone(timedelta(hours=-3))  # horário de Brasília

VERSAO = "2026-07-31.2"   # aparece no verificar.py e no cabeçalho do /admin

PASTA_STATIC = AQUI / "static"

# A pasta é criada se faltar, para o serviço subir e conseguir explicar o que
# falta em vez de estourar erro no import (que também derrubaria o verificar.py
# e o criar_usuario.py).
try:
    (PASTA_STATIC / "vendor").mkdir(parents=True, exist_ok=True)
except OSError:
    pass

FALTANDO = [
    nome for nome in ("login.html", "portal.html", "admin.html",
                      "admin-login.html", "vendor/chart.umd.js")
    if not (PASTA_STATIC / nome).exists()
]

app = FastAPI(title="Portal do Cliente — Logás", docs_url=None, redoc_url=None)
app.mount("/static", StaticFiles(directory=PASTA_STATIC), name="static")


def pagina(nome: str):
    """Serve uma tela ou, se o arquivo não foi copiado, diz exatamente o que falta."""
    arq = PASTA_STATIC / nome
    if arq.exists():
        return FileResponse(arq)
    return HTMLResponse(
        "<h1>Portal incompleto</h1><p>Falta o arquivo <code>static/"
        + nome + "</code> na pasta do portal. Copie a pasta <code>static</code> "
        "inteira (com a subpasta <code>vendor</code>) para junto do "
        "<code>app.py</code> e recarregue.</p>",
        status_code=500,
    )


# --------------------------------------------------------------------------- #
# Leitura do dados.js (com cache por mtime)
# --------------------------------------------------------------------------- #

_cache: dict = {"mtime": None, "dados": {"atualizado": None, "leituras": {}}}


def _js_para_json(txt: str) -> str:
    """
    O dados.js é objeto literal JavaScript, não JSON: as chaves vêm sem aspas
    (`t:`, `p:`, `obs:`). Esta varredura põe aspas nas chaves, remove comentários
    e vírgula sobrando, sem tocar no conteúdo das strings.
    """
    out: list[str] = []
    i, n = 0, len(txt)
    while i < n:
        c = txt[i]

        if c == '"':                                    # string: copia inteira
            j = i + 1
            while j < n:
                if txt[j] == "\\":
                    j += 2
                    continue
                if txt[j] == '"':
                    break
                j += 1
            out.append(txt[i : j + 1])
            i = j + 1
            continue

        if c == "/" and i + 1 < n and txt[i + 1] == "*":  # comentário /* */
            fim = txt.find("*/", i + 2)
            i = n if fim < 0 else fim + 2
            continue

        if c == "/" and i + 1 < n and txt[i + 1] == "/":  # comentário //
            fim = txt.find("\n", i)
            i = n if fim < 0 else fim
            continue

        if c.isascii() and (c.isalpha() or c == "_"):     # identificador
            j = i
            while j < n and (txt[j].isalnum() or txt[j] in "_$"):
                j += 1
            ident = txt[i:j]
            k = j
            while k < n and txt[k] in " \t\r\n":
                k += 1
            if k < n and txt[k] == ":" and ident not in ("true", "false", "null"):
                out.append(json.dumps(ident))            # chave -> "chave"
            else:
                out.append(ident)
            i = j
            continue

        if c == ",":                                      # vírgula antes de } ]
            k = i + 1
            while k < n and txt[k] in " \t\r\n":
                k += 1
            if k < n and txt[k] in "}]":
                i += 1
                continue

        out.append(c)
        i += 1

    return "".join(out)


def ler_dados() -> dict:
    """Extrai o objeto de window.LOGAS_DADOS = {...}; e devolve como dict."""
    arquivo = caminho_dados()
    try:
        if not arquivo.is_file():      # caminho aponta para pasta, ou não existe
            return _cache["dados"]
        mtime = arquivo.stat().st_mtime
    except OSError:
        return _cache["dados"]

    if _cache["mtime"] == mtime:
        return _cache["dados"]

    try:
        txt = arquivo.read_text(encoding="utf-8-sig", errors="replace")
    except OSError:                    # sem permissão, arquivo travado, etc.
        return _cache["dados"]
    ini = txt.find("=")
    ini = txt.find("{", ini if ini > 0 else 0)
    fim = txt.rfind("}")
    if ini < 0 or fim <= ini:
        return _cache["dados"]

    try:
        dados = json.loads(_js_para_json(txt[ini : fim + 1]))
    except json.JSONDecodeError:
        # Arquivo pela metade (o pipeline pode estar escrevendo agora):
        # mantém a última versão boa e tenta de novo na próxima requisição.
        return _cache["dados"]

    if not isinstance(dados.get("leituras"), dict):
        dados["leituras"] = {}
    dados.pop("_processados", None)          # lista de screenshots, não interessa aqui

    _cache.update(mtime=mtime, dados=dados)
    return dados


def ler_json(caminho: Path, padrao: dict) -> dict:
    # utf-8-sig: o PowerShell 5.1 grava JSON com BOM, e o json.loads engasga nele.
    try:
        return json.loads(caminho.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return padrao


def caminho_dados() -> Path:
    """Onde está o dados.js. Ordem: variável LOGAS_DADOS_JS, chave "dados_js" do
    config.json, arquivo dados.js nesta pasta. No Windows aceita tanto
    C:/painel/dados.js quanto C:\\painel\\dados.js."""
    do_ambiente = os.environ.get("LOGAS_DADOS_JS", "").strip()
    if do_ambiente:
        return Path(do_ambiente)
    do_config = str(config().get("dados_js", "")).strip()
    if do_config:
        return Path(do_config)
    return AQUI / "dados.js"


def admins() -> dict:
    return ler_json(ADMINS_JSON, {})


def salvar_json(caminho: Path, dados: dict) -> None:
    """Grava trocando o arquivo de uma vez, para nunca deixar um clientes.json
    pela metade se faltar energia no meio da escrita."""
    tmp = caminho.with_suffix(caminho.suffix + ".tmp")
    tmp.write_text(json.dumps(dados, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(str(tmp), str(caminho))
    try:
        caminho.chmod(0o600)
    except OSError:
        pass


ALFABETO_SENHA = "abcdefghjkmnpqrstuvwxyz23456789"   # sem 0/O/1/l/i, que confundem


def gerar_senha() -> str:
    """Senha em quatro blocos, fácil de ditar por telefone."""
    return "-".join(
        "".join(secrets.choice(ALFABETO_SENHA) for _ in range(4)) for _ in range(4)
    )


def nova_credencial(senha: str) -> dict:
    salt = secrets.token_hex(16)
    return {"salt": salt, "senha_hash": hash_senha(senha, salt)}


def registrar_log(quem: str, acao: str) -> None:
    """Histórico do que foi feito no painel administrativo."""
    try:
        with (AQUI / "admin.log").open("a", encoding="utf-8") as f:
            f.write(f"{datetime.now(TZ):%Y-%m-%d %H:%M:%S}\t{quem}\t{acao}\n")
    except OSError:
        pass


def pontos_disponiveis() -> list:
    """Pontos do dados.js, sem as leituras cujo nome o OCR não identificou e sem
    as que são auxiliares de outro ponto (temperatura, pressão de saída)."""
    auxiliares = pontos_auxiliares(config())
    return sorted(k for k in ler_dados().get("leituras", {})
                  if not k.startswith("?@") and k not in auxiliares)


LOGIN_VALIDO = re.compile(r"^[a-z0-9][a-z0-9._-]{2,31}$")


def config() -> dict:
    return ler_json(
        CONFIG_JSON,
        {"min_padrao": 30, "min": {}, "viagem_h": {}, "buffer_h": 1,
         "fallback_barh": 20, "max_bar": 260, "max_queda_barh": 100},
    )


def clientes() -> dict:
    return ler_json(CLIENTES_JSON, {})


# --------------------------------------------------------------------------- #
# Cálculos (mesmas regras da Torre de Controle)
# --------------------------------------------------------------------------- #


def _dt(s: str) -> datetime | None:
    try:
        d = datetime.fromisoformat(str(s).replace("Z", "+00:00"))
    except (TypeError, ValueError):
        return None
    return d.replace(tzinfo=TZ) if d.tzinfo is None else d


def iso_local(d: datetime | None) -> str | None:
    """ISO sem offset, no relógio de Brasília. O navegador do cliente interpreta
    string sem fuso como hora local, então o horário exibido é sempre o mesmo
    que foi registrado na foto, independente do fuso do aparelho dele."""
    if d is None:
        return None
    return d.astimezone(TZ).replace(tzinfo=None).isoformat(timespec="seconds")


def leituras(dados: dict, ponto: str) -> list[dict]:
    brutas = dados["leituras"].get(ponto) or []
    saida = []
    for r in brutas:
        t = _dt(r.get("t"))
        if t is None:
            continue
        saida.append({"t": t, "p": r.get("p"), "obs": r.get("obs")})
    saida.sort(key=lambda r: r["t"])

    # O pipeline às vezes grava a mesma leitura duas vezes no mesmo horário.
    # Duplicata faria a "leitura anterior" virar a própria leitura atual e
    # poluiria o gráfico, então fica só a última de cada horário.
    unicas = {}
    for r in saida:
        unicas[r["t"]] = r
    return [unicas[t] for t in sorted(unicas)]


def consumo_medido(ls: list, cfg: dict):
    """Média das últimas quedas reais em Bar/h. Subidas = troca de carreta."""
    taxas = [q / dh for _, _, q, dh in _quedas(ls, cfg)]
    if not taxas:
        return None
    ult = taxas[-4:]
    return sum(ult) / len(ult)


def ultima_valida(ls: list[dict], cfg: dict) -> dict | None:
    """Última leitura numérica, descartando salto impossível (erro de OCR)."""
    val = [r for r in ls if r["p"] is not None]
    if not val:
        return None
    ultima = val[-1]
    anterior = next((r for r in reversed(val[:-1]) if r["p"] <= cfg["max_bar"]), None)
    if ultima["p"] > cfg["max_bar"]:
        return anterior or ultima
    if anterior:
        dh = (ultima["t"] - anterior["t"]).total_seconds() / 3600
        if dh > 0 and (anterior["p"] - ultima["p"]) / dh > cfg["max_queda_barh"]:
            return anterior
    return ultima


def _quedas(ls: list, cfg: dict) -> list:
    """Pares de leituras consecutivas com queda real de pressão, já sem os saltos
    de troca de carreta e sem queda impossível (erro de leitura)."""
    val = [r for r in ls if r["p"] is not None and r["p"] <= cfg["max_bar"]]
    saida = []
    for ant, at in zip(val, val[1:]):
        dh = (at["t"] - ant["t"]).total_seconds() / 3600
        if dh <= 0:
            continue
        queda = ant["p"] - at["p"]
        if queda > 0 and queda / dh <= cfg["max_queda_barh"]:
            saida.append((ant, at, queda, dh))
    return saida


def consumo_por_dia(ls: list, cfg: dict, dias: int = 30) -> list:
    """Bar consumidos por dia. A queda entra no dia da leitura mais recente."""
    total = {}
    for _, at, queda, _ in _quedas(ls, cfg):
        d = at["t"].date()
        total[d] = total.get(d, 0.0) + queda
    corte = (datetime.now(TZ) - timedelta(days=dias)).date()
    return [{"dia": d.isoformat(), "bar": round(v, 1)}
            for d, v in sorted(total.items()) if d >= corte]


def consumo_janela(ls: list, cfg: dict, ini: datetime, fim: datetime) -> float:
    return round(sum(q for _, at, q, _ in _quedas(ls, cfg) if ini <= at["t"] < fim), 1)


def trocas_carreta(ls: list, cfg: dict) -> dict:
    """Subida forte de pressão entre duas fotos = carreta trocada."""
    salto = cfg.get("salto_troca", 25)
    val = [r for r in ls if r["p"] is not None and r["p"] <= cfg["max_bar"]]
    eventos = []
    for ant, at in zip(val, val[1:]):
        if at["p"] - ant["p"] >= salto:
            eventos.append({"t": at["t"], "de": round(ant["p"], 1), "para": round(at["p"], 1)})

    agora = datetime.now(TZ)
    intervalos = [
        (b["t"] - a["t"]).total_seconds() / 3600 for a, b in zip(eventos, eventos[1:])
    ]
    recentes = intervalos[-6:]
    return {
        "ultima": iso_local(eventos[-1]["t"]) if eventos else None,
        "horas_desde_ultima": (round((agora - eventos[-1]["t"]).total_seconds() / 3600, 1)
                               if eventos else None),
        "em_7_dias": sum(1 for e in eventos if (agora - e["t"]).days < 7),
        "em_30_dias": sum(1 for e in eventos if (agora - e["t"]).days < 30),
        "intervalo_medio_h": round(sum(recentes) / len(recentes), 1) if recentes else None,
        "eventos": [{"t": iso_local(e["t"]), "de": e["de"], "para": e["para"]}
                    for e in eventos[-10:]][::-1],
    }


def auxiliares_de(dados: dict, cfg: dict, ponto: str) -> list:
    """Leituras que acompanham um ponto sem serem tanque: temperatura, pressão
    de saída, o que mais o painel passar a ler. Não entram em consumo nem em
    autonomia; aparecem como tiles a mais no bloco do cliente.

    Configuração em config.json:
        "auxiliares": {
          "ORLÂNDIA": [
            {"ponto": "ORLÂNDIA TEMP",     "rotulo": "Temperatura",      "unidade": "C"},
            {"ponto": "ORLÂNDIA P.SAÍDA",  "rotulo": "Pressão de saída", "unidade": "Bar"}
          ]
        }
    """
    saida = []
    for item in cfg.get("auxiliares", {}).get(ponto, []) or []:
        nome = item.get("ponto")
        if not nome:
            continue
        ls = [r for r in leituras(dados, nome) if r["p"] is not None]
        if not ls:
            continue
        ultima = ls[-1]
        saida.append({
            "rotulo": item.get("rotulo") or nome,
            "valor": round(ultima["p"], 1),
            "unidade": item.get("unidade", ""),
            "casas": int(item.get("casas", 1)),
            "medido_em": iso_local(ultima["t"]),
        })
    return saida


def pontos_auxiliares(cfg: dict) -> set:
    """Nomes que são leitura auxiliar de algum ponto, e portanto não devem
    aparecer como ponto avulso no cadastro."""
    nomes = set()
    for lista in (cfg.get("auxiliares", {}) or {}).values():
        for item in lista or []:
            if item.get("ponto"):
                nomes.add(item["ponto"])
    return nomes


def estado_ponto(dados: dict, cfg: dict, ponto: str, horas_hist: int) -> dict:
    ls = leituras(dados, ponto)
    minimo = cfg["min"].get(ponto, cfg["min_padrao"])
    agora = datetime.now(TZ)

    corte = agora - timedelta(hours=horas_hist)
    historico = [
        {"t": iso_local(r["t"]), "p": r["p"]} for r in ls if r["t"] >= corte
    ]

    reg = ultima_valida(ls, cfg)
    if reg is None:
        return {"ponto": ponto, "pressao": None, "status": "sem_dados",
                "minimo": minimo, "historico": historico}

    sem_sinal = bool(ls) and ls[-1]["p"] is None
    taxa_medida = consumo_medido(ls, cfg)
    taxa = taxa_medida if taxa_medida is not None else cfg["fallback_barh"]

    # Fora do ar: projeta a queda desde a última leitura conhecida.
    horas_sem_leitura = (agora - reg["t"]).total_seconds() / 3600
    pressao, projetado = reg["p"], False
    if sem_sinal and taxa > 0 and horas_sem_leitura > 0:
        pressao = max(0.0, reg["p"] - taxa * horas_sem_leitura)
        projetado = True

    horas = (pressao - minimo) / taxa if taxa > 0 else None
    viagem = cfg["viagem_h"].get(ponto)
    chegada = None
    if viagem is not None and horas is not None:
        folga = max(0.0, horas - (viagem + cfg["buffer_h"]))
        chegada = iso_local(agora + timedelta(hours=folga + viagem))

    if horas is None:
        status = "ok"
    elif horas <= 0:
        status = "abaixo"
    elif horas < 2:
        status = "critico"
    elif horas < 12:
        status = "atencao"
    else:
        status = "ok"
    if sem_sinal:
        status = "sem_sinal" if horas is None or horas > 2 else status

    # Leitura anterior e o que caiu entre as duas ultimas fotos.
    val = [r for r in ls if r["p"] is not None and r["p"] <= cfg["max_bar"]]
    anterior = None
    ultima_variacao = None
    pos = next((i for i in range(len(val) - 1, -1, -1) if val[i]["t"] == reg["t"]),
               len(val) - 1)
    if pos >= 1:
        ant = val[pos - 1]
        dh = (reg["t"] - ant["t"]).total_seconds() / 3600
        anterior = {"p": round(ant["p"], 1), "t": iso_local(ant["t"])}
        if dh > 0:
            delta = ant["p"] - reg["p"]
            ultima_variacao = {
                "bar": round(abs(delta), 1),
                "horas": round(dh, 1),
                "barh": round(abs(delta) / dh, 2),
                "troca": delta < 0 and abs(delta) >= cfg.get("salto_troca", 25),
                "subiu": delta < 0,
            }

    m3_bar = cfg.get("m3_por_bar", {}).get(ponto)
    semana_atual = consumo_janela(ls, cfg, agora - timedelta(days=7), agora)
    semana_anterior = consumo_janela(ls, cfg, agora - timedelta(days=14),
                                     agora - timedelta(days=7))
    variacao = (round((semana_atual - semana_anterior) / semana_anterior * 100)
                if semana_anterior else None)
    dias = consumo_por_dia(ls, cfg, 30)

    return {
        "ponto": ponto,
        "pressao": round(pressao, 1),
        "pressao_lida": round(reg["p"], 1),
        "m3_por_bar": m3_bar,
        "auxiliares": auxiliares_de(dados, cfg, ponto),
        "anterior": anterior,
        "ultima_variacao": ultima_variacao,
        "leituras_registradas": len([r for r in ls if r["p"] is not None]),
        "por_dia": dias,
        "media_dia_bar": (round(sum(d["bar"] for d in dias[-7:]) / len(dias[-7:]), 1)
                          if dias else None),
        "trocas": trocas_carreta(ls, cfg),
        "semana": {"atual": semana_atual, "anterior": semana_anterior,
                   "variacao_pct": variacao},
        "medido_em": iso_local(reg["t"]),
        "projetado": projetado,
        "sem_sinal": sem_sinal,
        "obs": ls[-1]["obs"] if sem_sinal else None,
        "minimo": minimo,
        "consumo_barh": round(taxa, 2),
        "consumo_estimado": taxa_medida is None,
        "horas_ate_limite": None if horas is None else round(horas, 1),
        "limite_em": None if horas is None else iso_local(agora + timedelta(hours=max(0.0, horas))),
        "chegada_prevista": chegada,
        "status": status,
        "historico": historico,
    }


def painel_do_cliente(login: str, horas_hist: int) -> dict:
    reg = clientes().get(login) or {}
    cfg = config()
    dados = ler_dados()
    pontos = [estado_ponto(dados, cfg, p, horas_hist) for p in reg.get("pontos", [])]

    # Baias da mesma instalação (Extrema, Lagarto): quem supre é a de maior
    # pressão e o status do conjunto vale para todas.
    if reg.get("conjunto") and pontos:
        com_p = [p for p in pontos if p.get("pressao") is not None]
        if com_p:
            sup = max(com_p, key=lambda p: p["pressao"])
            taxas = [p["consumo_barh"] for p in com_p if p.get("consumo_barh")]
            for p in pontos:
                p["suprindo"] = p["ponto"] == sup["ponto"]
                p["status_conjunto"] = sup["status"]
                p["horas_conjunto"] = sup["horas_ate_limite"]
                p["chegada_conjunto"] = sup["chegada_prevista"]
                p["consumo_conjunto"] = max(taxas) if taxas else None

    return {
        "cliente": reg.get("nome", login),
        "versao": VERSAO,
        "conjunto": bool(reg.get("conjunto")),
        "unidade": "Bar",
        "atualizado": ler_dados().get("atualizado"),
        "horas_historico": horas_hist,
        "pontos": pontos,
    }


# --------------------------------------------------------------------------- #
# Sessão e autenticação
# --------------------------------------------------------------------------- #

COOKIE = "logas_portal"        # sessão de cliente
COOKIE_ADMIN = "logas_admin"   # sessão de administrador
_tentativas: dict[str, list[float]] = {}


def _b64e(b: bytes) -> str:
    return urlsafe_b64encode(b).decode().rstrip("=")


def _b64d(s: str) -> bytes:
    return urlsafe_b64decode(s + "=" * (-len(s) % 4))


def assinar(login: str, papel: str = "cliente") -> str:
    corpo = _b64e(json.dumps(
        {"u": login, "p": papel, "exp": time.time() + SESSAO_HORAS * 3600}
    ).encode())
    sig = _b64e(hmac.new(SECRET, corpo.encode(), hashlib.sha256).digest())
    return f"{corpo}.{sig}"


def validar(token: str | None, papel: str = "cliente") -> str | None:
    if not token or "." not in token:
        return None
    corpo, sig = token.rsplit(".", 1)
    esperado = _b64e(hmac.new(SECRET, corpo.encode(), hashlib.sha256).digest())
    if not hmac.compare_digest(sig, esperado):
        return None
    try:
        dados = json.loads(_b64d(corpo))
    except (ValueError, json.JSONDecodeError):
        return None
    if float(dados.get("exp", 0)) < time.time():
        return None
    if dados.get("p", "cliente") != papel:
        return None
    login = dados.get("u")
    if papel == "admin":
        return login if login in admins() else None
    reg = clientes().get(login)
    if not reg or reg.get("ativo") is False:   # acesso removido ou suspenso
        return None
    return login


def hash_senha(senha: str, salt: str) -> str:
    return hashlib.pbkdf2_hmac(
        "sha256", senha.encode(), bytes.fromhex(salt), 240_000
    ).hex()


def senha_confere(login: str, senha: str, papel: str = "cliente") -> bool:
    base = admins() if papel == "admin" else clientes()
    reg = base.get(login)
    if reg and reg.get("ativo") is False:
        return False
    if not reg or not reg.get("salt") or not reg.get("senha_hash"):
        # Compara mesmo assim, para não vazar por tempo de resposta se o login existe.
        hash_senha(senha, "00" * 16)
        return False
    return hmac.compare_digest(hash_senha(senha, reg["salt"]), reg["senha_hash"])


def bloqueado(chave: str) -> bool:
    agora = time.time()
    janela = [t for t in _tentativas.get(chave, []) if agora - t < 600]
    _tentativas[chave] = janela
    return len(janela) >= 8


def registrar_falha(chave: str) -> None:
    _tentativas.setdefault(chave, []).append(time.time())


def ip_visitante(request: Request) -> str:
    """IP real de quem acessou. Atrás de IIS ou de proxy, request.client.host é
    sempre 127.0.0.1, o que juntaria todos os clientes no mesmo balde do limite
    de tentativas de login. Com "confiar_proxy": true no config.json, o IP passa
    a vir dos cabeçalhos que o proxy escreve.

    Só ligue essa opção se o portal estiver mesmo atrás de um proxy seu: caso
    contrário qualquer visitante poderia inventar o cabeçalho e escapar do
    limite de tentativas.
    """
    if config().get("confiar_proxy"):
        cf = request.headers.get("cf-connecting-ip")
        if cf:
            return cf.strip()
        xff = request.headers.get("x-forwarded-for")
        if xff:
            return xff.split(",")[0].strip()
    return request.client.host if request.client else "?"


def usuario(request: Request) -> str | None:
    return validar(request.cookies.get(COOKIE))


def admin_atual(request: Request) -> str | None:
    return validar(request.cookies.get(COOKIE_ADMIN), papel="admin")


# --------------------------------------------------------------------------- #
# Rotas
# --------------------------------------------------------------------------- #


@app.get("/")
def raiz(request: Request):
    destino = "/painel" if usuario(request) else "/entrar"
    return RedirectResponse(destino, status_code=302)


@app.get("/entrar")
def tela_login(request: Request):
    if usuario(request):
        return RedirectResponse("/painel", status_code=302)
    return pagina("login.html")


@app.post("/entrar")
def entrar(request: Request, login: str = Form(...), senha: str = Form(...)):
    login = login.strip().lower()
    chave = f"{ip_visitante(request)}|{login}"
    if bloqueado(chave):
        return JSONResponse(
            {"erro": "Muitas tentativas. Espere 10 minutos ou fale com a Logás."},
            status_code=429,
        )
    if not senha_confere(login, senha):
        registrar_falha(chave)
        return JSONResponse({"erro": "Login ou senha incorretos."}, status_code=401)

    # Guarda o último acesso, para o painel administrativo mostrar quem está usando.
    try:
        base = clientes()
        if login in base:
            base[login]["ultimo_acesso"] = iso_local(datetime.now(TZ))
            salvar_json(CLIENTES_JSON, base)
    except OSError:
        pass

    resp = JSONResponse({"ok": True, "destino": "/painel"})
    resp.set_cookie(
        COOKIE, assinar(login), httponly=True, secure=COOKIE_SECURE,
        samesite="lax", max_age=int(SESSAO_HORAS * 3600), path="/",
    )
    return resp


@app.post("/sair")
def sair():
    resp = JSONResponse({"ok": True})
    resp.delete_cookie(COOKIE, path="/")
    return resp


@app.get("/painel")
def painel(request: Request):
    if not usuario(request):
        return RedirectResponse("/entrar", status_code=302)
    resp = pagina("portal.html")
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.get("/api/consumo")
def api_consumo(request: Request, horas: int = 72):
    login = usuario(request)
    if not login:
        return JSONResponse({"erro": "sessao_expirada"}, status_code=401)
    horas = max(6, min(int(horas), 24 * 30))
    resp = JSONResponse(painel_do_cliente(login, horas))
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.get("/api/excel")
def api_excel(request: Request, horas: int = 168):
    login = usuario(request)
    if not login:
        return JSONResponse({"erro": "sessao_expirada"}, status_code=401)
    horas = max(6, min(int(horas), 24 * 30))
    dados = painel_do_cliente(login, horas)

    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ARIAL = "Arial"
    cab_fonte = Font(name=ARIAL, bold=True, color="FFFFFF")
    cab_fundo = PatternFill("solid", fgColor="13203A")

    def escreve(ws, cabecalho, linhas, larguras, formatos=None):
        ws.append(cabecalho)
        for c in range(1, len(cabecalho) + 1):
            cel = ws.cell(row=1, column=c)
            cel.font, cel.fill = cab_fonte, cab_fundo
            cel.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(c)].width = larguras[c - 1]
        for lin in linhas:
            ws.append(lin)
        for row in ws.iter_rows(min_row=2):
            for cel in row:
                cel.font = Font(name=ARIAL)
                if formatos and formatos[cel.column - 1]:
                    cel.number_format = formatos[cel.column - 1]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb = Workbook()

    def dt_local(iso):
        d = _dt(iso)
        return d.replace(tzinfo=None) if d else None

    ws = wb.active
    ws.title = "Resumo"
    conj = dados["conjunto"]
    ROT = {"ok": "Normal", "atencao": "Atenção", "critico": "Crítico",
           "abaixo": "Abaixo da reserva", "sem_sinal": "Sem sinal",
           "sem_dados": "Sem leitura"}
    linhas = []
    for p in dados["pontos"]:
        h = p.get("horas_conjunto") if conj else p.get("horas_ate_limite")
        cheg = p.get("chegada_conjunto") if conj else p.get("chegada_prevista")
        st = p.get("status_conjunto") if conj else p.get("status")
        lin = [
            p["ponto"], p.get("pressao"), p.get("minimo"),
            p.get("consumo_conjunto") or p.get("consumo_barh"),
            p.get("media_dia_bar"), h,
            dt_local(p.get("limite_em")), dt_local(cheg),
            dt_local((p.get("trocas") or {}).get("ultima")),
            dt_local(p.get("medido_em")), ROT.get(st, ""),
        ]
        if conj:
            lin.append("Sim" if p.get("suprindo") else "Não")
        linhas.append(lin)

    cabecalho = ["Ponto", "Pressão (Bar)", "Reserva mínima (Bar)",
                 "Consumo da instalação (Bar/h)" if conj else "Consumo (Bar/h)",
                 "Média por dia (Bar)", "Autonomia (h)", "Chega na reserva",
                 "Reabastecimento previsto", "Última troca de carreta",
                 "Medido em", "Situação"]
    larguras = [22, 13, 18, 16, 16, 12, 20, 22, 22, 18, 18]
    formatos = [None, "0.00", "0.00", "0.00", "0.0", "0.0", "dd/mm/yyyy hh:mm",
                "dd/mm/yyyy hh:mm", "dd/mm/yyyy hh:mm", "dd/mm/yyyy hh:mm", None]
    if conj:
        cabecalho.append("Baia em uso")
        larguras.append(12)
        formatos.append(None)
    escreve(ws, cabecalho, linhas, larguras, formatos)

    ws2 = wb.create_sheet("Leituras")
    lin2 = [
        [p["ponto"], dt_local(r["t"]), r["p"]]
        for p in dados["pontos"] for r in p.get("historico", [])
    ]
    escreve(ws2, ["Ponto", "Data e hora", "Pressão (Bar)"], lin2,
            [22, 20, 14], [None, "dd/mm/yyyy hh:mm", "0.00"])

    ws4 = wb.create_sheet("Consumo por dia")
    lin4 = [[p["ponto"], _dt(d["dia"]).replace(tzinfo=None) if _dt(d["dia"]) else d["dia"],
             d["bar"], (round(d["bar"] * p["m3_por_bar"], 1) if p.get("m3_por_bar") else None)]
            for p in dados["pontos"] for d in (p.get("por_dia") or [])]
    escreve(ws4, ["Ponto", "Dia", "Consumo (Bar)", "Consumo (m³)"], lin4,
            [22, 14, 15, 15], [None, "dd/mm/yyyy", "0.0", "0.0"])

    ws5 = wb.create_sheet("Trocas de carreta")
    lin5 = [[p["ponto"], dt_local(e["t"]), e["de"], e["para"]]
            for p in dados["pontos"] for e in ((p.get("trocas") or {}).get("eventos") or [])]
    escreve(ws5, ["Ponto", "Data e hora", "Antes (Bar)", "Depois (Bar)"], lin5,
            [22, 20, 14, 14], [None, "dd/mm/yyyy hh:mm", "0.0", "0.0"])

    ws3 = wb.create_sheet("Informações")
    ws3.append(["Cliente", dados["cliente"]])
    ws3.append(["Período do histórico", f"últimas {horas} horas"])
    ws3.append(["Gerado em", datetime.now(TZ).replace(tzinfo=None)])
    ws3.append(["Última leitura registrada", dt_local(dados["atualizado"])])
    ws3.append([])
    ws3.append(["Autonomia e reabastecimento são estimativas calculadas pelo consumo médio"])
    ws3.append(["recente e podem variar conforme a operação."])
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 26
    for row in ws3.iter_rows():
        for cel in row:
            cel.font = Font(name=ARIAL)
            if isinstance(cel.value, datetime):
                cel.number_format = "dd/mm/yyyy hh:mm"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug = re.sub(r"[^a-z0-9]+", "-", dados["cliente"].lower()).strip("-")
    nome = f"consumo-gnc-{slug}-{datetime.now(TZ):%Y-%m-%d}.xlsx"
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome}"',
                 "Cache-Control": "no-store"},
    )


@app.middleware("http")
async def cabecalhos(request: Request, call_next):
    resp: Response = await call_next(request)
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["Referrer-Policy"] = "no-referrer"
    resp.headers["X-Frame-Options"] = "DENY"
    return resp


# --------------------------------------------------------------------------- #
# Painel administrativo
# --------------------------------------------------------------------------- #


@app.get("/favicon.ico")
def favicon():
    return Response(status_code=204)


@app.get("/admin")
def admin_home(request: Request):
    if not admin_atual(request):
        return RedirectResponse("/admin/entrar", status_code=302)
    resp = pagina("admin.html")
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.get("/admin/entrar")
def admin_tela(request: Request):
    if admin_atual(request):
        return RedirectResponse("/admin", status_code=302)
    if not admins():
        return HTMLResponse(
            "<h1>Nenhum administrador cadastrado</h1><p>No servidor, rode "
            "<code>python criar_admin.py</code> para criar o primeiro acesso "
            "administrativo. Depois recarregue esta página.</p>",
            status_code=503,
        )
    return pagina("admin-login.html")


@app.post("/admin/entrar")
def admin_login(request: Request, login: str = Form(...), senha: str = Form(...)):
    login = login.strip().lower()
    ip = ip_visitante(request)
    chave = f"admin|{ip}|{login}"
    if bloqueado(chave):
        return JSONResponse(
            {"erro": "Muitas tentativas. Espere 10 minutos."}, status_code=429
        )
    if not senha_confere(login, senha, papel="admin"):
        registrar_falha(chave)
        registrar_log(login, f"tentativa de login administrativo falhou (IP {ip})")
        return JSONResponse({"erro": "Login ou senha incorretos."}, status_code=401)

    registrar_log(login, f"entrou no painel administrativo (IP {ip})")
    resp = JSONResponse({"ok": True, "destino": "/admin"})
    resp.set_cookie(
        COOKIE_ADMIN, assinar(login, papel="admin"), httponly=True,
        secure=COOKIE_SECURE, samesite="lax",
        max_age=int(SESSAO_HORAS * 3600), path="/",
    )
    return resp


@app.post("/admin/sair")
def admin_sair():
    resp = JSONResponse({"ok": True})
    resp.delete_cookie(COOKIE_ADMIN, path="/")
    return resp


def _exige_admin(request: Request):
    return admin_atual(request)


@app.get("/admin/api/estado")
def admin_estado(request: Request):
    quem = _exige_admin(request)
    if not quem:
        return JSONResponse({"erro": "sessao_expirada"}, status_code=401)

    disponiveis = pontos_disponiveis()
    usados = {}
    for login, reg in clientes().items():
        for p in reg.get("pontos", []):
            usados.setdefault(p, []).append(login)

    lista = []
    for login, reg in sorted(clientes().items()):
        pontos = reg.get("pontos", [])
        lista.append({
            "login": login,
            "nome": reg.get("nome", login),
            "pontos": pontos,
            "conjunto": bool(reg.get("conjunto")),
            "ativo": reg.get("ativo") is not False,
            "criado_em": reg.get("criado_em"),
            "senha_trocada_em": reg.get("senha_trocada_em"),
            "ultimo_acesso": reg.get("ultimo_acesso"),
            "inexistentes": [p for p in pontos if p not in disponiveis],
        })

    return JSONResponse({
        "admin": quem,
        "versao": VERSAO,
        "pontos": disponiveis,
        "pontos_sem_cliente": [p for p in disponiveis if p not in usados],
        "clientes": lista,
        "atualizado": ler_dados().get("atualizado"),
    })


@app.post("/admin/api/cliente")
async def admin_salvar_cliente(request: Request):
    quem = _exige_admin(request)
    if not quem:
        return JSONResponse({"erro": "sessao_expirada"}, status_code=401)

    corpo = await request.json()
    login = str(corpo.get("login", "")).strip().lower()
    nome = str(corpo.get("nome", "")).strip()
    pontos = [str(p) for p in (corpo.get("pontos") or [])]
    conjunto = bool(corpo.get("conjunto"))

    if not LOGIN_VALIDO.match(login):
        return JSONResponse(
            {"erro": "Login: de 3 a 32 caracteres, só letras minúsculas, números, "
                     "ponto, hífen ou sublinhado."}, status_code=400)
    if not nome:
        return JSONResponse({"erro": "Informe o nome que aparece na tela."}, status_code=400)
    if not pontos:
        return JSONResponse({"erro": "Selecione ao menos um ponto."}, status_code=400)

    disponiveis = pontos_disponiveis()
    faltando = [p for p in pontos if p not in disponiveis]
    if faltando:
        return JSONResponse(
            {"erro": "Não existe no dados.js: " + ", ".join(faltando)}, status_code=400)

    base = clientes()
    novo = login not in base
    reg = dict(base.get(login, {}))
    reg.update({"nome": nome, "pontos": pontos, "conjunto": conjunto})
    reg.setdefault("ativo", True)

    senha = None
    if novo:
        senha = gerar_senha()
        reg.update(nova_credencial(senha))
        reg["criado_em"] = iso_local(datetime.now(TZ))
        reg["senha_trocada_em"] = reg["criado_em"]

    base[login] = reg
    salvar_json(CLIENTES_JSON, base)
    registrar_log(quem, ("criou" if novo else "editou") + f" o cliente {login} "
                        f"({', '.join(pontos)})")

    return JSONResponse({"ok": True, "criado": novo, "senha": senha})


@app.post("/admin/api/senha")
async def admin_nova_senha(request: Request):
    quem = _exige_admin(request)
    if not quem:
        return JSONResponse({"erro": "sessao_expirada"}, status_code=401)

    login = str((await request.json()).get("login", "")).strip().lower()
    base = clientes()
    if login not in base:
        return JSONResponse({"erro": "Cliente não encontrado."}, status_code=404)

    senha = gerar_senha()
    base[login].update(nova_credencial(senha))
    base[login]["senha_trocada_em"] = iso_local(datetime.now(TZ))
    salvar_json(CLIENTES_JSON, base)
    registrar_log(quem, f"gerou nova senha para {login}")
    return JSONResponse({"ok": True, "senha": senha})


@app.post("/admin/api/situacao")
async def admin_situacao(request: Request):
    quem = _exige_admin(request)
    if not quem:
        return JSONResponse({"erro": "sessao_expirada"}, status_code=401)

    corpo = await request.json()
    login = str(corpo.get("login", "")).strip().lower()
    ativo = bool(corpo.get("ativo"))
    base = clientes()
    if login not in base:
        return JSONResponse({"erro": "Cliente não encontrado."}, status_code=404)

    base[login]["ativo"] = ativo
    salvar_json(CLIENTES_JSON, base)
    registrar_log(quem, ("reativou" if ativo else "suspendeu") + f" o acesso de {login}")
    return JSONResponse({"ok": True})


@app.post("/admin/api/remover")
async def admin_remover(request: Request):
    quem = _exige_admin(request)
    if not quem:
        return JSONResponse({"erro": "sessao_expirada"}, status_code=401)

    login = str((await request.json()).get("login", "")).strip().lower()
    base = clientes()
    if login not in base:
        return JSONResponse({"erro": "Cliente não encontrado."}, status_code=404)

    base.pop(login)
    salvar_json(CLIENTES_JSON, base)
    registrar_log(quem, f"removeu o cliente {login}")
    return JSONResponse({"ok": True})


@app.get("/admin/api/historico")
def admin_historico(request: Request, linhas: int = 60):
    if not _exige_admin(request):
        return JSONResponse({"erro": "sessao_expirada"}, status_code=401)
    arq = AQUI / "admin.log"
    if not arq.exists():
        return JSONResponse({"linhas": []})
    todas = arq.read_text(encoding="utf-8", errors="replace").splitlines()
    return JSONResponse({"linhas": todas[-max(10, min(linhas, 500)):][::-1]})
