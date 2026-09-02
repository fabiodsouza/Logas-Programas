#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ler_painel.py - Leitor OCR do painel "Centro de Operacoes Logistica" (Logas).
Detecta cada manometro (capsula cinza), le o NOME acima e o VALOR/status abaixo,
le a data/hora e grava no dados.js (window.LOGAS_DADOS).

Uso:
  python ler_painel.py foto.png
  python ler_painel.py pasta/            (processa todas as imagens, em ordem)
  python ler_painel.py foto.png --dry-run
  python ler_painel.py foto.png --csv saida.csv
"""
import sys, os, re, json, glob, argparse, unicodedata, datetime
import cv2, numpy as np, pytesseract, difflib
import shutil as _shutil
# No Windows, se o tesseract nao estiver no PATH, procura no local padrao de instalacao
if os.name=="nt" and not _shutil.which("tesseract"):
    for _p in (r"C:\Program Files\Tesseract-OCR\tesseract.exe",
               r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"):
        if os.path.exists(_p):
            pytesseract.pytesseract.tesseract_cmd=_p; break

# ----------------------------- cadastro -----------------------------
ROSTER = {
    "BETIM": ["ALSCO","BELGO","EXTRUMINAS","M.PEQUI","VDL","MG STEEL","VILMA",
              "G.PERDÕES","JD CANADÁ","FORTLEV","COMPRESSOR BETIM"],
    "POÇOS DE CALDAS": ["ACQUION"],
    "POUSO ALEGRE": ["BALANÇA","POUSO"],
    "EXTREMA-MG": ["BAIA 1","BAIA 2","BAIA 3"],
    "IPATINGA": ["FERMAG","PLANALTO","TROPICAL","BARRIGÃO"],
    "ESPÍRITO SANTO": ["SPINASSÉ","COLATINA"],
    "BARBACENA": ["AMG SP","DOW"],
    "LAGARTO-SE": ["BAIA 1","BAIA 2","SERGÁS"],
    "SÃO PAULO": ["ORLÂNDIA"],
    "MATO GROSSO DO SUL": ["SEARA"],
}
REGIOES = list(ROSTER.keys())

# Empresas ainda em implantacao: sempre gravadas como "Em Implantacao"
# (ignora o OCR do valor). Remova daqui quando entrarem em operacao.
EM_IMPLANTACAO = {"SEARA"}

# Queda de pressao fisicamente impossivel numa hora => provavel erro de OCR.
# Acima disso (Bar/h) a leitura e marcada "conferir" em vez de virar alarme falso.
MAX_QUEDA_BARH = 100
# Pressao maxima plausivel: o manometro vai ate ~250 Bar; acima disso e erro de OCR.
MAX_BAR = 260

# Carimbo de versao: vai para a 1a linha do dados.js. Serve para saber qual
# script gerou o arquivo (ex.: quando a rotina agendada aponta para outra copia).
VERSAO = "ler_painel.py v2 - com EVENTOS CRITICOS"

# ---- caixa "EVENTOS CRITICOS" (rodape) ----
# Estes itens NAO tem capsula/manometro: sao pares "rotulo  valor" dentro de um
# retangulo de borda vermelha no rodape. Por isso tem um caminho de leitura
# proprio (detect_caixa_eventos + ler_eventos) e limites proprios de sanidade,
# ja que nao estao em Bar de manometro.
# (rotulo esperado, chave no dados.js, unidade, minimo, maximo)
EVENTOS = [
    ("TEMPERATURA ORLANDIA",   "ORLÂNDIA TEMP",    "°C", -20, 120),
    ("PRESSAO SAIDA ORLANDIA", "ORLÂNDIA P.SAÍDA", "Bar",  0, 400),
]
EVENTO_POR_CHAVE = {k:(u,lo,hi) for _,k,u,lo,hi in EVENTOS}

def chave(regiao, empresa):
    if regiao == "LAGARTO-SE" and empresa in ("BAIA 1","BAIA 2"):
        return empresa + " (LAGARTO)"
    return empresa

STATUS_PHRASES = ["SEM SINAL","SEM ENERGIA","AGUARDANDO STARLINK","EM IMPLANTACAO"]
STATUS_LABEL = {"SEM SINAL":"Sem Sinal","SEM ENERGIA":"Sem Energia",
                "AGUARDANDO STARLINK":"Aguardando Starlink","EM IMPLANTACAO":"Em Implantação"}
STATUS_KEY={"SEM SINAL":"SINAL","SEM ENERGIA":"ENERGIA","AGUARDANDO STARLINK":"AGUARDANDO","EM IMPLANTACAO":"IMPLANTACAO"}

def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD",s) if unicodedata.category(c)!="Mn")
def norm(s):
    return re.sub(r"[^A-Z0-9 .-]","",strip_accents(s).upper()).strip()

NAME_INDEX=[(norm(e),reg,e) for reg,emps in ROSTER.items() for e in emps]
REG_INDEX=[(norm(r),r) for r in REGIOES]

# ----------------------------- visao -----------------------------
def detect_caps(img):
    H,W=img.shape[:2]
    hsv=cv2.cvtColor(img,cv2.COLOR_BGR2HSV); S,V=hsv[:,:,1],hsv[:,:,2]
    mask=((S<60)&(V>120)&(V<235)).astype("uint8")*255
    mask=cv2.morphologyEx(mask,cv2.MORPH_CLOSE,np.ones((9,9),np.uint8))
    n,lab,stats,_=cv2.connectedComponentsWithStats(mask,8)
    cand=[]
    for i in range(1,n):
        x,y,w,h,a=stats[i]
        if h<=w or a<0.40*w*h: continue
        if w<0.015*W or w>0.14*W: continue
        if h<0.04*H or h>0.34*H: continue
        if not (1.2<h/w<3.3): continue
        cand.append([int(x),int(y),int(w),int(h)])
    if not cand: return [],0,0
    mw=float(np.median([c[2] for c in cand])); mh=float(np.median([c[3] for c in cand]))
    caps=[c for c in cand if 0.6*mw<c[2]<1.6*mw and 0.6*mh<c[3]<1.6*mh]
    caps.sort(key=lambda c:(round(c[1]/(mh*0.6)),c[0]))
    return caps, mw, mh

def detect_headers(img):
    hsv=cv2.cvtColor(img,cv2.COLOR_BGR2HSV)
    H_,S_,V_=hsv[:,:,0],hsv[:,:,1],hsv[:,:,2]
    cyan=((H_>80)&(H_<105)&(S_>80)&(V_>120)).astype("uint8")*255
    inv=255-cv2.dilate(cyan,np.ones((2,2),np.uint8))
    d=pytesseract.image_to_data(inv,lang="eng",config="--psm 11",output_type=pytesseract.Output.DICT)
    toks=[]
    for i,t in enumerate(d["text"]):
        t=t.strip()
        if not t: continue
        try: c=int(float(d["conf"][i]))
        except: c=-1
        if c<20 or len(norm(t))<2: continue
        toks.append(dict(cx=d["left"][i]+d["width"][i]//2, cy=d["top"][i]+d["height"][i]//2,
                         h=d["height"][i], text=t))
    # agrupa tokens por linha e casa com nomes de regiao
    toks.sort(key=lambda t:(round(t["cy"]/max(8,np.median([x["h"] for x in toks]) if toks else 8)),t["cx"]))
    headers=[]; i=0
    while i<len(toks):
        grp=[toks[i]]; j=i+1
        while j<len(toks) and abs(toks[j]["cy"]-toks[i]["cy"])<0.8*toks[i]["h"] and toks[j]["cx"]-grp[-1]["cx"]<6*toks[i]["h"]:
            grp.append(toks[j]); j+=1
        cand=norm(" ".join(g["text"] for g in grp))
        best=None; br=0
        for rn,r in REG_INDEX:
            ra=difflib.SequenceMatcher(None,cand,rn).ratio()
            if rn in cand or cand in rn: ra=max(ra,0.9)
            if ra>br: br=ra; best=r
        if best and br>=0.6:
            cx=sum(g["cx"] for g in grp)//len(grp); cy=min(g["cy"] for g in grp)
            headers.append((best,cx,cy))
        i=j
    return headers

def _ocr_linhas(bin_img):
    """OCR de uma faixa, agrupando as palavras por linha.
    Retorna [(texto, conf)]. Combina psm 6 e psm 11 porque em faixas largas e
    esparsas um pega o que o outro perde."""
    saida=[]
    for psm in (6,11):
        try:
            d=pytesseract.image_to_data(bin_img,lang="eng",config="--psm %d"%psm,
                                        output_type=pytesseract.Output.DICT)
        except Exception:
            continue
        linhas={}
        for i,t in enumerate(d["text"]):
            t=t.strip()
            if not t: continue
            try: c=max(int(float(d["conf"][i])),0)
            except Exception: c=0
            # agrupa por faixa vertical (psm 11 nao encadeia linhas de forma confiavel)
            yc=d["top"][i]+d["height"][i]//2
            achou=None
            for k in linhas:
                if abs(k-yc)<=max(8,d["height"][i]//2): achou=k; break
            box=(d["left"][i],d["top"][i],d["width"][i],d["height"][i])
            linhas.setdefault(achou if achou is not None else yc,[]).append((d["left"][i],t,c,box))
        for _,itens in sorted(linhas.items()):
            itens.sort()
            texto=" ".join(t for _,t,_,_ in itens)
            conf=int(np.mean([c for _,t,c,_ in itens]))
            # bbox do ultimo token numerico da linha: e nele que o valor sera
            # relido com a votacao de digitos (best_num), bem mais confiavel
            # bbox da regiao do valor. 1a escolha: token que e um numero limpo.
            # 2a: token que ao menos contem digito ("2..CO", "1l,00"). 3a: ultimo
            # token da linha, desde que separado do anterior por um vao grande,
            # que e como o painel diagrama rotulo <espaco> valor.
            bb=None
            for _,t,_,box in itens:
                if re.fullmatch(r"\d{1,4}(?:[.,]\d{1,2})?",t.strip()): bb=box
            if bb is None:
                for _,t,_,box in itens:
                    if re.search(r"\d",t): bb=box
            if bb is None and len(itens)>=3:
                lx,lt,lc,lbox=itens[-1]; px,pt,pc,pbox=itens[-2]
                alt=max(lbox[3],1)
                if lbox[0]-(pbox[0]+pbox[2])>=1.5*alt: bb=lbox
            saida.append((texto,conf,bb))
    return saida

def _casa_evento(rotulo):
    """Casa o rotulo lido com o cadastro EVENTOS. Alem da similaridade global,
    aceita match por palavras-chave, que sobrevive melhor a erro de OCR
    ('Pressdo Saida Orlandia', 'Temperatura Criandia')."""
    melhor=None; br=0.0
    for lbl,k,u,lo,hi in EVENTOS:
        r=difflib.SequenceMatcher(None,rotulo,lbl).ratio()
        if lbl in rotulo or rotulo in lbl: r=max(r,0.93)
        toks=[t for t in lbl.split() if len(t)>3]
        if toks:
            ok=sum(1 for t in toks
                   if any(difflib.SequenceMatcher(None,w,t).ratio()>=0.75 for w in rotulo.split()))
            if ok==len(toks): r=max(r,0.90)
            elif ok>=len(toks)-1 and len(toks)>=2: r=max(r,0.72)
        if r>br: br=r; melhor=(k,u,lo,hi)
    return melhor,br

def _bins_rodape(crop):
    """Variantes binarizadas da faixa de eventos.
    O texto do rodape e fino e sem contorno; uma unica passada com Otsu costuma
    quebrar as letras ('Temperatura' vira 'Tenpecrstana'). Aqui o crop e ampliado
    bastante e sao geradas versoes com limiar de Otsu, limiar fixo e com o traco
    engrossado, para o OCR votar entre elas."""
    g=cv2.cvtColor(crop,cv2.COLOR_BGR2GRAY)
    mx=int(g.max()) or 255
    saida=[]
    for f in (4.0,6.0):
        up=cv2.resize(g,None,fx=f,fy=f,interpolation=cv2.INTER_CUBIC)
        up=cv2.GaussianBlur(up,(3,3),0)
        otsu=cv2.threshold(up,0,255,cv2.THRESH_BINARY_INV+cv2.THRESH_OTSU)[1]
        fixo=cv2.threshold(up,max(40,int(mx*0.42)),255,cv2.THRESH_BINARY_INV)[1]
        for b in (otsu,fixo):
            saida.append(b)
            # engrossa o traco: o texto esta escuro sobre fundo claro, entao
            # dilatar o escuro = erodir a imagem
            saida.append(cv2.erode(b,np.ones((2,2),np.uint8),iterations=1))
    return saida,4.0

def ler_eventos(img, y_rodape=None, debug_dir=None, tag=None, verbose=False):
    """Le os pares 'rotulo  valor' do bloco EVENTOS CRITICOS (rodape).
    Nao depende de moldura nem do titulo: varre a faixa abaixo dos manometros."""
    H,W=img.shape[:2]
    y0=int(y_rodape) if y_rodape else int(0.80*H)
    y0=max(0,min(y0,H-20)); x0=int(0.14*W)
    crop=img[y0:H, x0:W].copy()
    if crop.size==0: return {}
    hsv=cv2.cvtColor(crop,cv2.COLOR_BGR2HSV)
    cyan=(((hsv[:,:,0]>80)&(hsv[:,:,0]<105)&(hsv[:,:,1]>80)&(hsv[:,:,2]>120))).astype("uint8")
    crop[cv2.dilate(cyan,np.ones((3,3),np.uint8)).astype(bool)]=(0,0,0)
    bins,fdbg=_bins_rodape(crop)
    if debug_dir:
        for i,b in enumerate(bins[:2]):
            try: cv2.imwrite(os.path.join(debug_dir,"debug_EVENTOS_%s_%d.png"%(tag or "crop",i)),b)
            except Exception: pass
    cand={}; brutas=[]; numeradas=[]
    for bi,b in enumerate(bins):
        f=4.0 if bi<4 else 6.0
        for texto,conf,bbnum in _ocr_linhas(b):
            if len(texto.strip())<6: continue
            brutas.append(texto)
            m=re.search(r"(\d{1,4})(?:[.,](\d{1,2}))?\s*$",texto.strip())
            rotulo=norm(texto[:m.start()] if m else texto)
            if len(rotulo)<6: continue
            if bbnum: numeradas.append((bbnum[1]/f, bbnum, f, rotulo))
            vtxt=None
            if m:
                try: vtxt=float(m.group(1)+("."+m.group(2) if m.group(2) else ""))
                except Exception: vtxt=None
            alvo,r=_casa_evento(rotulo)
            if not alvo or r<0.62: continue
            # guarda varios candidatos por chave: se o recorte do numero do
            # melhor deles nao render leitura, tenta o seguinte
            cand.setdefault(alvo[0],[]).append((r,conf,bbnum,f,rotulo,alvo,vtxt))
    # Preenchimento por posicao: o painel lista os itens sempre na mesma ordem
    # (temperatura, depois pressao de saida). Se o OCR do rotulo degradou a ponto
    # de nao casar, mas o numero de linhas com valor bate com o cadastro, as
    # chaves que faltam sao atribuidas pela ordem vertical. Vao marcadas com
    # obs "conferir (rotulo ilegivel)" para nao passarem por leitura confirmada.
    if len(cand)<len(EVENTOS) and numeradas:
        linhas=[]
        for y,bb,f,rot in sorted(numeradas):
            if linhas and abs(linhas[-1][0]-y)<=6: continue
            linhas.append((y,bb,f,rot))
        if len(linhas)==len(EVENTOS):
            for (y,bb,f,rot),(lbl,k,u,lo,hi) in zip(linhas,EVENTOS):
                if k in cand: continue
                cand[k]=[(0.0,0,bb,f,rot,(k,u,lo,hi),None)]
            if verbose:
                print("  [eventos] rotulo ilegivel em %d item(ns); atribuido pela ordem do painel"
                      % sum(1 for v in cand.values() if v[0][0]==0.0))
    out={}
    for k,lst in cand.items():
        lst.sort(key=lambda c:(-c[0],-c[1]))
        r,conf,_,_,rotulo,alvo,_=lst[0]
        _k,u,lo,hi=alvo
        st=_match_status(rotulo)
        if st: out[k]=(None,st,80); continue
        # Votacao do valor: junta as releituras do recorte do numero (caminho
        # best_num, com whitelist de digitos) com os numeros que sairam no texto
        # de cada binarizacao. Vence o mais votado; empate, o de maior confianca.
        # Region do valor: cada binarizacao devolve um bbox para o mesmo numero,
        # alguns cortados (pegam so o "1" de "21"). Em vez de votar entre eles, os
        # bboxes sao levados para a coordenada original, os outliers descartados
        # pela mediana, e a UNIAO dos restantes vira o recorte lido. Assim o
        # numero entra inteiro. As leituras individuais e os numeros que sairam
        # no texto entram so como voto de apoio, com peso menor.
        cxs=[]
        for r2,c2,bbnum,f,_rot,_a,vtxt in lst:
            if not bbnum: continue
            bx,by,bw,bh=[int(v/f) for v in bbnum]
            if bw<4 or bh<4: continue
            cxs.append((bx,by,bw,bh))
        votos={}
        def _voto(v,c,peso):
            if v is None: return
            a,b=votos.get(v,(0,0)); votos[v]=(a+peso*max(c,1),max(b,c))
        if cxs:
            mcy=sorted(by+bh//2 for _,by,_,bh in cxs)[len(cxs)//2]
            malt=sorted(bh for _,_,_,bh in cxs)[len(cxs)//2]
            iguais=[c for c in cxs if abs(c[1]+c[3]//2-mcy)<=max(malt,6)]
            if iguais:
                ux=min(c[0] for c in iguais); uy=min(c[1] for c in iguais)
                ux2=max(c[0]+c[2] for c in iguais); uy2=max(c[1]+c[3] for c in iguais)
                mx_=max(3,int((ux2-ux)*0.15)); my=max(3,int((uy2-uy)*0.35))
                sub=img[max(0,y0+uy-my):min(H,y0+uy2+my), max(0,x0+ux-mx_):min(W,x0+ux2+mx_)]
                if sub.size:
                    v2,o2,c3=ler_valor(sub); _voto(v2,c3,8)
                for c in iguais[:6]:
                    bx,by,bw,bh=c
                    mx_=max(3,int(bw*0.15)); my=max(3,int(bh*0.35))
                    sub=img[max(0,y0+by-my):min(H,y0+by+bh+my), max(0,x0+bx-mx_):min(W,x0+bx+bw+mx_)]
                    if sub.size:
                        v2,o2,c3=ler_valor(sub); _voto(v2,c3,1)
        for r2,c2,bbnum,f,_rot,_a,vtxt in lst:
            _voto(vtxt,c2,1)
        val=None
        if votos:
            val=max(votos.items(),key=lambda kv:(kv[1][0],kv[1][1]))[0]
            conf=max(conf,votos[val][1])
        if val is None: continue
        obs=None if r>=0.62 else "conferir (rotulo ilegivel)"
        if not (lo<=val<=hi):
            obs="conferir (%s fora de %g a %g %s; leu %g)"%(k,lo,hi,u,val); val=None
        out[k]=(val,obs,conf)
    if verbose and not out:
        print("  [eventos] nada casou no rodape (y>=%d). Linhas lidas: %s"
              % (y0, "; ".join(v[:60] for v in brutas[:8]) or "(nenhuma)"))
    return out

# ----------------------------- leitura de texto -----------------------------
def cor_predominante(bgr):
    hsv=cv2.cvtColor(bgr,cv2.COLOR_BGR2HSV); S,V=hsv[:,:,1],hsv[:,:,2]
    fg=V>110
    if fg.sum()<10: return "vazio"
    sat=S[fg].mean()
    if sat<55: return "cinza"
    hue=hsv[:,:,0][fg&(S>80)]
    if len(hue):
        hm=np.median(hue)
        if hm<15 or hm>165: return "vermelho"
        if 35<hm<90: return "verde"
        return "laranja"
    return "cinza"

def _ocr(img,psm,wl=None):
    cfg="--psm %d"%psm
    if wl: cfg+=" -c tessedit_char_whitelist="+wl
    return pytesseract.image_to_data(img,lang="eng",config=cfg,output_type=pytesseract.Output.DICT)

def ler_nome(crop):
    if crop.size==0: return None,0
    up=cv2.resize(crop,None,fx=3,fy=3,interpolation=cv2.INTER_CUBIC)
    g=cv2.cvtColor(up,cv2.COLOR_BGR2GRAY)
    g=cv2.threshold(g,0,255,cv2.THRESH_BINARY_INV+cv2.THRESH_OTSU)[1]
    txt=pytesseract.image_to_string(g,lang="eng",
        config="--psm 7 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚÃÕÇ. 0123456789").strip()
    cand=norm(txt)
    if len(cand)<2: return None,0
    best=None; br=0
    for nm,reg,emp in NAME_INDEX:
        r=difflib.SequenceMatcher(None,cand,nm).ratio()
        if cand==nm: r=1.0
        elif len(cand)>=4 and (cand in nm or nm in cand) and abs(len(cand)-len(nm))<=2: r=max(r,0.93)
        if r>br: br=r; best=(reg,emp)
    return (best, int(br*100)) if br>=0.62 else (None,int(br*100))

def _bins(crop):
    """Varias binarizacoes do valor para votacao. Inclui mascaras de COR alem do
    cinza: 'colored' pega qualquer texto saturado (verde/laranja/vermelho) e
    'warm' isola so as cores quentes (laranja/vermelho/amarelo). Isolar o digito
    pela cor reduz a confusao tipica dos numeros laranja (ex.: 5 lido como 9),
    porque tira o fundo escuro e as bordas da capsula da jogada.
    A 1a binarizacao (gray_inv) e a usada para detectar status em texto."""
    up=cv2.resize(crop,None,fx=5,fy=5,interpolation=cv2.INTER_CUBIC)
    upb=cv2.GaussianBlur(up,(3,3),0)              # suaviza serrilhado (anti-aliasing)
    g=cv2.cvtColor(up,cv2.COLOR_BGR2GRAY)         # nitido para status/Otsu
    hsv=cv2.cvtColor(upb,cv2.COLOR_BGR2HSV)       # suavizado para as mascaras de cor
    S,V,Hh=hsv[:,:,1],hsv[:,:,2],hsv[:,:,0]
    gray_inv=cv2.threshold(g,0,255,cv2.THRESH_BINARY_INV+cv2.THRESH_OTSU)[1]
    colored =255-(((S>55)&(V>80)).astype("uint8")*255)
    warm    =255-((((Hh<35)|(Hh>160))&(S>60)&(V>80)).astype("uint8")*255)
    return [gray_inv, colored, warm]

def _match_status(nt):
    for ph,key in STATUS_KEY.items():
        if key in nt: return STATUS_LABEL[ph]
        for w in nt.split():
            if len(w)>=5 and difflib.SequenceMatcher(None,w,key).ratio()>=0.8: return STATUS_LABEL[ph]
    return None

def best_num(ims):
    """Le o numero em todas as binarizacoes/psm e VOTA: soma a confianca de cada
    valor lido (com bonus para quem traz casas decimais, ex.: 56,05). O valor mais
    'votado' vence. Assim, uma unica leitura errada de alta confianca (97) nao
    ganha de varias leituras coerentes do valor certo (57)."""
    from collections import defaultdict
    score=defaultdict(float); maxc=defaultdict(int)
    for im in ims:
        for psm in (7,6,8,13):
            d=_ocr(im,psm,"0123456789,.")
            for i,t in enumerate(d["text"]):
                t=t.strip().replace(".",",")
                m=re.fullmatch(r"(\d{1,3})(?:,(\d{1,2}))?",t)
                if not m: continue
                try: c=int(float(d["conf"][i]))
                except: c=0
                if c<0: c=0
                dec=1 if m.group(2) else 0
                val=float(m.group(1)+("."+m.group(2) if m.group(2) else ""))
                key=(round(val,2),dec)
                # bonus por decimais e por nº de digitos (prefere 182,97 a um "2" solto)
                score[key]+=c+(20 if dec else 0)+(len(m.group(1))-1)*4
                if c>maxc[key]: maxc[key]=c
    if not score: return None
    val,dec=max(score, key=lambda k:score[k])
    return (dec, maxc[(val,dec)], val)

def _recorta_numero(crop):
    """Mantem so a 1a faixa de texto (o numero), descartando o 'Bar' abaixo do
    manometro — que distorcia o limiar e fazia o 5 virar 9."""
    if crop is None or crop.size==0: return crop
    import numpy as _np
    g=cv2.cvtColor(crop,cv2.COLOR_BGR2GRAY); mx=int(g.max())
    if mx<40: return crop
    fg=g>max(55,int(mx*0.45))
    rows=fg.sum(axis=1).astype(float)
    if rows.max()<1: return crop
    on=rows>max(2.0, rows.max()*0.18)
    n=len(on); i=0
    while i<n and not on[i]: i+=1
    a=i
    while i<n and on[i]: i+=1
    b=i
    if b-a>=6:
        pad=3
        return crop[max(0,a-pad):min(crop.shape[0],b+pad), :]
    return crop

def ler_valor(crop):
    if crop.size==0: return None,"conferir",0
    crop=_recorta_numero(crop)
    cor=cor_predominante(crop); bins=_bins(crop)
    for psm in (6,7):
        txt=pytesseract.image_to_string(bins[0],lang="eng",config="--psm %d"%psm).strip()
        st=_match_status(norm(txt))
        if st: return None,st,80
    b=best_num(bins)
    if b: return b[2],None,b[1]
    if cor=="vermelho": return None,"Sem Sinal",40
    return None,"conferir",0

# ----------------------------- pipeline -----------------------------
def regiao_da_capsula(cap, headers, mw):
    cx=cap[0]+cap[2]//2; cy=cap[1]
    cands=[(cy-hy, r) for (r,hx,hy) in headers if hy<cy and abs(hx-cx)<3.2*mw]
    if cands: return min(cands)[1]
    cands=[(cy-hy, r) for (r,hx,hy) in headers if hy<cy]
    if cands: return min(cands)[1]
    return None

def ts_do_nome(path):
    """Extrai data/hora do nome do arquivo telemetria_AAAA-MM-DD_HH-MM-SS."""
    m=re.search(r"(\d{4})-(\d{2})-(\d{2})[_ ](\d{2})-(\d{2})-(\d{2})",os.path.basename(path))
    if m:
        y,mo,d,H,Mi,S=m.groups()
        return "%s-%s-%sT%s:%s:%s"%(y,mo,d,H,Mi,S)
    return None

def processar_imagem(path, debug=None):
    raw=cv2.imread(path)
    if raw is None: raise RuntimeError("Nao consegui abrir a imagem: "+path)
    H,W=raw.shape[:2]; scale=max(1.0,1700.0/W)
    img=cv2.resize(raw,None,fx=scale,fy=scale,interpolation=cv2.INTER_CUBIC) if scale>1.05 else raw
    H,W=img.shape[:2]
    ts=ts_do_nome(path) or ler_timestamp(img) or datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    caps,mw,mh=detect_caps(img)
    lidas=[]   # (cx, y, reg_nome, emp, p, obs, vcf, ncf)
    for cap in caps:
        x,y,w,h=cap; cx=x+w//2; nb=int(w*1.35)
        ncrop=img[max(0,int(y-0.42*h)):max(0,int(y-0.02*h)), max(0,cx-nb):min(W,cx+nb)]
        vcrop=img[min(H,y+h):min(H,int(y+h+0.45*h)), max(0,cx-nb):min(W,cx+nb)]
        match,ncf=ler_nome(ncrop)
        p,obs,vcf=ler_valor(vcrop)
        reg=match[0] if match else None; emp=match[1] if match else None
        if debug and emp and norm(emp)==norm(debug):
            _bn=os.path.splitext(os.path.basename(path))[0]; _tag=debug.replace(" ","_").replace("/","_")
            _dir=os.path.dirname(os.path.abspath(__file__))
            try:
                cv2.imwrite(os.path.join(_dir,"debug_%s_%s_nome.png"%(_tag,_bn)), ncrop)
                cv2.imwrite(os.path.join(_dir,"debug_%s_%s_valor.png"%(_tag,_bn)), vcrop)
                print("  [debug] %s @ (%d,%d): nome_conf=%s | valor lido='%s' valor_conf=%s -> recortes salvos (debug_%s_*.png)"
                      %(emp,cx,y,ncf,(obs if p is None else "%.2f"%p),vcf,_tag))
            except Exception as _e:
                print("  [debug] falha ao salvar recortes: %s"%_e)
        if emp in EM_IMPLANTACAO:
            p=None; obs="Em Implantação"; vcf=100
        lidas.append([cx,y,reg,emp,p,obs,vcf,ncf])

    # --- BAIAs: decidir EXTREMA (esq) vs LAGARTO (dir) por posicao X e ordenar ---
    baias=[r for r in lidas if r[3] in ("BAIA 1","BAIA 2","BAIA 3")]
    if baias:
        xs=sorted(b[0] for b in baias)
        corte=None
        if len(xs)>=2:
            gaps=[(xs[i+1]-xs[i],i) for i in range(len(xs)-1)]
            g,i=max(gaps)
            if g>2.5*mw: corte=(xs[i]+xs[i+1])/2
        for b in baias:
            esq = (corte is None) or (b[0]<corte)
            b[2]="EXTREMA-MG" if esq else "LAGARTO-SE"
        for grupo,reg in (([b for b in baias if b[2]=="EXTREMA-MG"],"EXTREMA-MG"),
                          ([b for b in baias if b[2]=="LAGARTO-SE"],"LAGARTO-SE")):
            for idx,b in enumerate(sorted(grupo,key=lambda r:r[0]),1):
                b[3]="BAIA %d"%idx

    leituras={}
    for cx,y,reg,emp,p,obs,vcf,ncf in lidas:
        if emp is None:
            emp="?@%d,%d"%(cx,y); obs=(obs+" | nome ilegivel") if obs else "nome ilegivel"; reg=reg or "?"
        if p is not None and vcf<40 and not obs:
            obs="conferir"
        k=chave(reg or "?",emp)
        if k not in leituras or vcf>leituras[k]["conf"]:
            leituras[k]=dict(regiao=reg or "?",empresa=emp,p=p,obs=obs,conf=vcf,ncf=ncf)

    # --- bloco EVENTOS CRITICOS (sem capsula: caminho de leitura proprio) ---
    # o rodape comeca logo abaixo do manometro mais baixo da tela
    _yr=max((y+h for x,y,w,h in caps), default=int(0.80*H))+int(0.55*(mh or 40))
    _dbg=os.path.dirname(os.path.abspath(__file__)) if (debug and norm(debug)=="EVENTOS") else None
    for k,(p,obs,cf) in ler_eventos(img,_yr,_dbg,os.path.splitext(os.path.basename(path))[0],
                                    verbose=True).items():
        leituras[k]=dict(regiao="EVENTOS",empresa=k,p=p,obs=obs,conf=cf,ncf=100)
    return ts,leituras

def ler_timestamp(img):
    H,W=img.shape[:2]
    crop=img[0:int(H*0.12),0:int(W*0.55)]
    crop=cv2.resize(crop,None,fx=2,fy=2,interpolation=cv2.INTER_CUBIC)
    txt=pytesseract.image_to_string(crop,lang="eng",
        config="--psm 6 -c tessedit_char_whitelist=0123456789:/")
    hh=re.search(r"(\d{1,2}:\d{2}:\d{2})",txt); dd=re.search(r"(\d{2}/\d{2}/\d{4})",txt)
    if dd and hh:
        d,m,y=dd.group(1).split("/"); H_,M_,S_=hh.group(1).split(":")
        return "%s-%s-%sT%02d:%s:%s"%(y,m,d,int(H_),M_,S_)
    return None

# ----------------------------- dados.js -----------------------------
DADOS_JS=os.path.join(os.path.dirname(os.path.abspath(__file__)),"dados.js")
XLSX_PADRAO=os.path.join(os.path.dirname(os.path.abspath(__file__)),"leituras_pivo.xlsx")
def carregar_dados_js(path=DADOS_JS):
    if not os.path.exists(path): return dict(atualizado=None,_processados=[],leituras={})
    txt=open(path,encoding="utf-8").read()
    m=re.search(r"window\.LOGAS_DADOS\s*=\s*(\{.*\})\s*;",txt,re.S)
    if not m: raise RuntimeError("Nao encontrei window.LOGAS_DADOS em dados.js")
    js=re.sub(r"([{,]\s*)([A-Za-z_][A-Za-z0-9_]*)\s*:",r'\1"\2":',m.group(1))
    js=re.sub(r",(\s*[}\]])",r"\1",js)
    return json.loads(js)
def fmt_p(p):
    if p is None: return "null"
    return str(int(p)) if float(p).is_integer() else "%.2f"%p
def escrever_dados_js(data,path=DADOS_JS):
    if os.path.exists(path):
        import shutil; shutil.copy2(path, path+".bak")
    L=["/* DADOS DE LEITURA - Painel Logas (gerado por %s) */"%VERSAO,
       "window.LOGAS_DADOS = {",'  atualizado: "%s",'%data["atualizado"],
       "  _processados: "+json.dumps(data["_processados"],ensure_ascii=False)+",","  leituras: {"]
    keys=list(data["leituras"].keys())
    for i,k in enumerate(keys):
        it=[]
        for r in data["leituras"][k]:
            s='{t:"%s", p:%s'%(r["t"],fmt_p(r.get("p")))
            if r.get("obs"): s+=', obs:"%s"'%r["obs"]
            it.append(s+"}")
        L.append("    %s: [ %s ]%s"%(json.dumps(k,ensure_ascii=False),", ".join(it),"," if i<len(keys)-1 else ""))
    L+=["  }","};",""]; open(path,"w",encoding="utf-8").write("\n".join(L))
def merge(data,ts,leituras,foto):
    if foto not in (data.get("_processados") or []): data.setdefault("_processados",[]).append(foto)
    data["atualizado"]=ts; lt=data.setdefault("leituras",{})
    for k,r in leituras.items():
        arr=lt.setdefault(k,[]); item={"t":ts,"p":r["p"]}
        if r["obs"]: item["obs"]=r["obs"]
        # Eventos criticos (temperatura / pressao de saida) nao passam pelas travas
        # de Bar do manometro: tem faixa propria, ja conferida em ler_eventos().
        if k in EVENTO_POR_CHAVE:
            pass
        # Trava de valor MAXIMO: acima do manometro (~250 Bar) e erro de OCR (digitos a mais).
        elif r["p"] is not None and r["p"] > MAX_BAR:
            print("  [ALERTA] %s: leitura %.0f Bar acima do maximo (%d) -> marcado p/ conferir" % (k, r["p"], MAX_BAR))
            item = {"t": ts, "p": None, "obs": "conferir (acima de %d Bar; leu %.0f)" % (MAX_BAR, r["p"])}
        # Trava anti-erro de OCR: queda absurda vs ultima leitura numerica = suspeita.
        # (Aumentos sao abastecimento/troca de carreta e sao mantidos.)
        elif r["p"] is not None and arr:
            prev=next((a for a in reversed(arr) if a.get("p") is not None and a["p"]<=MAX_BAR), None)  # ignora valores-lixo (>MAX_BAR) como base
            if prev is not None:
                try:
                    dh=(datetime.datetime.fromisoformat(ts)-datetime.datetime.fromisoformat(prev["t"])).total_seconds()/3600.0
                except Exception:
                    dh=0
                queda=prev["p"]-r["p"]
                if dh>0 and queda>0 and queda/dh>MAX_QUEDA_BARH:
                    print("  [ALERTA] %s: queda suspeita %.2f -> %.2f em %.1fh (%.0f Bar/h) -> marcado p/ conferir"
                          %(k,prev["p"],r["p"],dh,queda/dh))
                    item={"t":ts,"p":None,"obs":"conferir (leu ~%.0f; anterior %.0f)"%(r["p"],prev["p"])}
        if arr and arr[-1].get("t")==ts: arr[-1]=item
        else: arr.append(item)
    return data

# ----------------------------- planilha pivo (.xlsx) -----------------------------
def _cab_coluna(k):
    """Cabecalho amigavel da coluna a partir da chave do dados.js."""
    if k in EVENTO_POR_CHAVE:             return "%s (%s)"%(k,EVENTO_POR_CHAVE[k][0])
    if k in ("BAIA 1","BAIA 2","BAIA 3"): return "Extrema "+k.replace("BAIA ","B")
    if k.endswith("(LAGARTO)"):           return "Lagarto B"+k.split()[1]
    return k

def _ordem_colunas(leituras):
    """Colunas na ordem do ROSTER (via chave); chaves extras vao para o fim."""
    ordem=[]
    for reg,emps in ROSTER.items():
        for e in emps:
            k=chave(reg,e)
            if k in leituras and k not in ordem: ordem.append(k)
    for k in leituras:
        if k not in ordem: ordem.append(k)
    return ordem

def gerar_xlsx(data, path=XLSX_PADRAO):
    """Gera um PIVO: 1 linha por horario, 1 coluna por empresa, celula = pressao.
    Sem cores (a formatacao fica a cargo do analista). Mantem o historico inteiro."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [xlsx] openpyxl nao instalado -> rode: pip install openpyxl  (planilha nao gerada)")
        return False
    leituras=data.get("leituras",{}) or {}
    if not leituras:
        print("  [xlsx] sem leituras para exportar."); return False
    ordem=_ordem_colunas(leituras)
    # uniao de todos os horarios + mapa (coluna,horario)->leitura
    tss=set(); val={}
    for k in ordem:
        m={}
        for r in leituras[k]:
            m[r["t"]]=r            # se repetir o horario, fica a ultima
            tss.add(r["t"])
        val[k]=m
    tss=sorted(tss)
    wb=Workbook(); ws=wb.active; ws.title="Leituras"
    ws.cell(row=1,column=1,value="Data")
    # por empresa: coluna do bar + coluna "#" (consumo), igual as abas de mes do modelo
    barcol={}; col=2
    for k in ordem:
        ws.cell(row=1,column=col,value=_cab_coluna(k)); barcol[k]=col
        if k in EVENTO_POR_CHAVE:   # temperatura / pressao de saida nao tem "consumo"
            col+=1; continue
        ws.cell(row=1,column=col+1,value="#")
        col+=2
    ncols=col-1
    for i,ts in enumerate(tss,start=2):
        c=ws.cell(row=i,column=1)
        try:
            c.value=datetime.datetime.fromisoformat(ts); c.number_format="dd/mm/yyyy hh:mm"
        except Exception:
            c.value=ts
        for k in ordem:
            bc=barcol[k]; r=val[k].get(ts)
            if r and r.get("p") is not None:
                ws.cell(row=i,column=bc,value=r["p"])    # status/sem sinal -> celula em branco
            if i>2 and k not in EVENTO_POR_CHAVE:   # consumo (#) = bar(hora anterior) - bar(hora atual); em branco se faltar leitura
                L=get_column_letter(bc)
                ws.cell(row=i,column=bc+1,
                        value='=IF(OR(%s%d="",%s%d=""),"",%s%d-%s%d)'%(L,i-1,L,i,L,i-1,L,i))
    # estetica simples: cabecalho em negrito, congela 1a linha/coluna, larguras
    bold=Font(bold=True)
    for j in range(1,ncols+1):
        cc=ws.cell(row=1,column=j); cc.font=bold; cc.alignment=Alignment(horizontal="center")
    ws.freeze_panes="B2"
    ws.column_dimensions["A"].width=17
    for k in ordem:
        bc=barcol[k]
        ws.column_dimensions[get_column_letter(bc)].width=11
        if k in EVENTO_POR_CHAVE: continue
        ws.column_dimensions[get_column_letter(bc+1)].width=7
    try:
        wb.save(path); return True
    except PermissionError:
        print("  [xlsx] nao consegui salvar (arquivo aberto no Excel?):",path); return False

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("entrada"); ap.add_argument("--dry-run",action="store_true"); ap.add_argument("--csv")
    ap.add_argument("--desde",help="processa so prints a partir desta data/hora, ex: 2026-06-24T07:00:00")
    ap.add_argument("--debug",help="salva os recortes (nome/valor) que o OCR le da empresa indicada, ex: --debug TROPICAL")
    ap.add_argument("--xlsx",nargs="?",const=XLSX_PADRAO,default=None,
                    help="gera tambem a planilha pivo .xlsx (padrao: leituras_pivo.xlsx ao lado do dados.js)")
    a=ap.parse_args()
    eh_pasta=os.path.isdir(a.entrada)
    if eh_pasta:
        # varre a pasta E todas as subpastas (estrutura por dia/hora),
        # processando apenas prints no padrao telemetria_*.png
        pats=("telemetria_*.png","telemetria_*.jpg","telemetria_*.PNG","telemetria_*.JPG")
        fotos=sorted(sum([glob.glob(os.path.join(a.entrada,"**",p),recursive=True) for p in pats],[]))
        if not fotos:  # se nao houver no padrao, cai para todas as imagens (recursivo)
            fotos=sorted(sum([glob.glob(os.path.join(a.entrada,"**",e),recursive=True) for e in ("*.png","*.jpg")],[]))
    else:
        fotos=[a.entrada]
    if not fotos: print("Nenhuma imagem."); return
    if a.desde:
        corte=a.desde.strip().replace(" ","T")
        antes=len(fotos)
        fotos=[f for f in fotos if (ts_do_nome(f) is None) or (ts_do_nome(f) >= corte)]
        print("filtro --desde %s: %d de %d prints no periodo."%(corte,len(fotos),antes))
    data=carregar_dados_js(); rows=[]
    # em modo pasta, ignora prints ja processados (so processa os novos)
    if eh_pasta:
        ja=set(data.get("_processados") or [])
        novos=[f for f in fotos if os.path.basename(f) not in ja]
        if not novos:
            print("Nenhum print novo (todos ja processados)."); return
        fotos=novos
    print("%d print(s) a processar."%len(fotos))
    for foto in fotos:
        ts,le=processar_imagem(foto, a.debug); print("\n=== %s  (%s) ==="%(os.path.basename(foto),ts)); fl=0
        for k in sorted(le):
            r=le[k]; val=(r["obs"] if r["p"] is None else "%.2f"%r["p"]) or "—"
            mk="  <-- CONFERIR" if (r["obs"] and "conferir" in r["obs"].lower()) or r["p"] is None and r["obs"] not in STATUS_LABEL.values() else ""
            if mk: fl+=1
            print("  %-22s %-12s %-22s (val %s / nome %s)%s"%(k,r["regiao"],val,r["conf"],r.get("ncf",0),mk))
            rows.append((r["empresa"],r["regiao"],ts,"" if r["p"] is None else ("%.2f"%r["p"]).replace(".",","),r["obs"] or "OK"))
        print("  %d lidas · %d p/ conferir"%(len(le),fl))
        if not a.dry_run: data=merge(data,ts,le,os.path.basename(foto))
    if not a.dry_run:
        escrever_dados_js(data); print("\ndados.js atualizado")
        xpath=a.xlsx or XLSX_PADRAO          # por padrao gera sempre, junto com o CSV/dados.js
        if gerar_xlsx(data,xpath): print("planilha pivo:",xpath)
    else:
        print("\n[dry-run]")
        if a.xlsx and gerar_xlsx(data,a.xlsx): print("planilha pivo:",a.xlsx)
    if a.csv:
        import csv
        with open(a.csv,"w",newline="",encoding="utf-8-sig") as f:
            w=csv.writer(f,delimiter=";"); w.writerow(["Empresa","Regiao","DataHora","PressaoBar","Status"]); w.writerows(rows)
        print("CSV:",a.csv)

if __name__=="__main__": main()
